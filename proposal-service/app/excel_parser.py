"""
Excel proposal parser.

Reads a Blue Lime proposal Excel workbook and produces a normalized
`ProposalData` dictionary that the generator can render.

The parser is tolerant of:
  - Missing cells (returns sensible defaults like "Not Included")
  - Placeholder values like "XXXX", "XXXXX", "N/A" → treated as absent
  - Trailing whitespace and minor capitalization variation
  - The "Kieth/Keith" typo in the current templates (left as-is — the parser
    doesn't try to autocorrect client-provided names)

If the Excel template structure changes, this is the file that needs to be
updated. Cell coordinates are documented inline so the mapping is easy to
follow.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PLACEHOLDER_VALUES = {"xxxx", "xxxxx", "xxxxxx", "n/a", "na", "0", "", "-"}


# -----------------------------------------------------------------------------
# Cell-reading utilities
# -----------------------------------------------------------------------------
def _cell(ws: Worksheet, coord: str) -> Any:
    """Read a cell, return None if empty."""
    v = ws[coord].value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _is_placeholder(v: Any) -> bool:
    """True if the value should be treated as 'not included'."""
    if v is None:
        return True
    if isinstance(v, (int, float)):
        return False  # 0 is a real number; treat as absent only via labels
    s = str(v).strip().lower()
    return s in PLACEHOLDER_VALUES


def _money(v: Any) -> Optional[float]:
    """Coerce a money-like cell to a float, or None if it can't be parsed."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _money_str(v: Any, default: str = "Not Included") -> str:
    """Format a money-like cell as '$X,XXX.XX' or fall back."""
    f = _money(v)
    if f is None or _is_placeholder(v):
        return default
    if f == int(f):
        return f"${int(f):,}"
    return f"${f:,.2f}"


def _limit_str(v: Any, default: str = "Not Included") -> str:
    """Format a limit cell — keep strings intact, format numbers with commas."""
    if _is_placeholder(v):
        return default
    if isinstance(v, (int, float)):
        return f"${int(v):,}" if v == int(v) else f"${v:,.2f}"
    return str(v).strip()


def _date_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%B %-d, %Y") if hasattr(v, "strftime") else str(v)
    return str(v)


# -----------------------------------------------------------------------------
# Section parsers
# -----------------------------------------------------------------------------
def _parse_client(wb) -> dict:
    """Read client info from the Summary tab (with Cover Page as backup)."""
    summary = wb["Summary"]
    cover = wb["Cover Page"] if "Cover Page" in wb.sheetnames else None

    name = _cell(summary, "D3") or (_cell(cover, "A23") if cover else None)
    address = _cell(summary, "D10") or ""
    eff = _cell(summary, "D14")
    exp = _cell(summary, "F14")

    name = (name or "Unnamed Association").strip()

    # short_name = the name minus a trailing "Owners Association, Inc." for headers
    short_name = re.sub(r"\s*Owners Association.*$", "", name, flags=re.IGNORECASE).strip()
    if not short_name:
        short_name = name

    eff_str = _date_str(eff) or "TBD"
    exp_str = _date_str(exp) or "TBD"
    policy_term = f"{eff_str} \u2013 {exp_str}"

    return {
        "name": name,
        "short_name": short_name,
        "address": address.strip(),
        "eff_date": eff_str,
        "exp_date": exp_str,
        "policy_term": policy_term,
        # homes/units typically isn't in Summary — could be added if templates
        # add a cell for it. For now leave a generic line.
        "homes": "",
        "homes_summary": "Per association records",
    }


def _parse_account_manager(wb) -> dict:
    summary = wb["Summary"]
    name = _cell(summary, "I7") or "David Ritualo"
    email = _cell(summary, "I9") or "DRitualo@bluelimeins.com"
    phone = _cell(summary, "I11") or "(210) 507-0262"

    # Map the AM name to a slug for the team-page badge.
    SLUG_MAP = {
        "briana howard":  "briana",
        "carol marquez":  "carol",
        "steven melgosa": "steven",
        "susan finke":    "susan",
        "david ritualo":  "david",
        "valerie cordes": "valerie",
        "daniel infante": "daniel",
    }
    slug = SLUG_MAP.get(name.strip().lower(), "david")

    return {
        "name": name,
        "title": "Account Manager",
        "email": email,
        "phone": phone,
        "slug": slug,
    }


def _parse_premium_comparison(wb) -> dict:
    """Read the Premium Summary tab — side-by-side expiring vs. proposed."""
    ps = wb["Premium Summary"]

    expiring_carrier = _cell(ps, "D12") or "Expiring"
    proposed_carrier = _cell(ps, "E12") or "Proposed"

    # Map of (label, expiring_coord, proposed_coord)
    rows = [
        ("General Liability",          "D14", "E14"),
        ("Total Property Limit",       "D15", "E15"),
        ("All Other Peril Deductible", "D19", "E19"),
        ("Wind / Hail Deductible",     "D20", "E20"),
        ("W/H Deductible BuyBack",     "D22", "E22"),
        ("Hired & Non-Owned Auto",     "D23", "E23"),
        ("Crime",                      "D24", "E24"),
        ("Directors & Officers",       "D25", "E25"),
        ("Cyber Liability",            "D26", "E26"),
        ("Umbrella – Excess Liability", "D28", "E28"),
        ("Volunteer Accident",         "D29", "E29"),
        ("Workers' Compensation",      "D30", "E30"),
    ]

    comparison_lines = []
    for label, exp_c, prop_c in rows:
        exp_v = _cell(ps, exp_c)
        prop_v = _cell(ps, prop_c)

        # Skip rows where both sides are placeholder/empty AND the row is
        # auxiliary (Cyber, Crime). This keeps the comparison table tight.
        if _is_placeholder(exp_v) and _is_placeholder(prop_v):
            if label in {"Cyber Liability", "W/H Deductible BuyBack", "Crime"}:
                continue

        comparison_lines.append({
            "label": label,
            "expiring": _limit_str(exp_v, "Not Included"),
            "proposed": _limit_str(prop_v, "Not Included"),
        })

    expiring_total = _money(_cell(ps, "D31")) or 0.0
    proposed_total = _money(_cell(ps, "E31")) or 0.0

    change = proposed_total - expiring_total
    if abs(change) < 1:
        change_note = "Proposed program substantially matches expiring at renewal."
    elif change > 0:
        change_note = ("Increase reflects updated property values, expanded "
                       "coverage parts, or carrier rate adjustments.")
    else:
        change_note = "Decrease reflects rate adjustments and program optimization."

    return {
        "expiring_carrier": expiring_carrier,
        "proposed_carrier": proposed_carrier,
        "comparison_lines": comparison_lines,
        "expiring_total": expiring_total,
        "proposed_total": proposed_total,
        "expiring_total_str": f"${expiring_total:,.2f}",
        "proposed_total_str": f"${proposed_total:,.2f}",
        "change_note": change_note,
    }


def _coverage_status(carrier_cell, limit_cells: list, default_carrier: str = "Philadelphia Insurance"):
    """Decide whether a coverage is included, and what carrier label to show.

    A coverage is "not included" only if the carrier is placeholder AND every
    limit cell is also placeholder/zero. If the carrier is missing but limits
    are populated, we infer the carrier from the proposal's primary carrier
    (default "Philadelphia Insurance") rather than dropping the whole
    coverage block.

    Returns a tuple (carrier_label, not_included).
    """
    carrier_is_placeholder = _is_placeholder(carrier_cell)
    has_real_limits = any(
        not _is_placeholder(c) and _money(c) not in (None, 0)
        for c in limit_cells
    )

    if not carrier_is_placeholder:
        return (str(carrier_cell).strip(), False)
    if has_real_limits:
        # Carrier missing, but limits are real → coverage IS included
        return (default_carrier, False)
    return ("Not Included", True)


def _parse_coverages(wb) -> list[dict]:
    """Read each coverage block from the Summary tab.

    Each coverage has a known starting row in the template. If a carrier shows
    "N/A" we mark it as not included only when the limits are also empty;
    otherwise we infer the carrier from the proposal's primary carrier.
    """
    s = wb["Summary"]

    # Identify the proposal's primary carrier — used as the fallback when a
    # specific coverage block has limits but no carrier name filled in.
    # Falls through several known-populated carrier cells in priority order.
    primary_carrier = None
    for coord in ("A26", "A43", "A50", "A72"):  # Property, GL, Auto, Umbrella
        v = _cell(s, coord)
        if not _is_placeholder(v):
            primary_carrier = str(v).strip()
            break
    primary_carrier = primary_carrier or "Philadelphia Insurance"

    coverages = []

    # Commercial Property — rows 25-30
    prop_carrier = _cell(s, "A26") or "Not Included"
    coverages.append({
        "title": "Commercial Property",
        "carrier": prop_carrier if not _is_placeholder(prop_carrier) else "Not Included",
        "not_included": _is_placeholder(prop_carrier),
        "description": (
            "Commercial property insurance helps pay to repair or replace "
            "buildings and other property damaged or destroyed by covered "
            "perils. This proposal covers property at replacement cost. Flood "
            "and Earthquake are excluded unless specifically listed."
        ),
        "panel_rows": [
            [
                ("Buildings",            _money_str(_cell(s, "C26"), "$0")),
                ("Business Personal Property", _money_str(_cell(s, "D26"), "$0")),
                ("Outdoor Property",     _money_str(_cell(s, "E26"), "$0")),
                ("Total Property Limit", _money_str(_cell(s, "G26"), "$0")),
            ],
            [
                ("AOP Deductible",       _money_str(_cell(s, "H26"), "Standard")),
                ("Wind / Hail Ded.",     _money_str(_cell(s, "I26"), "Standard")),
                ("Equip. Breakdown Ded.", "Included"),
                ("Named Storm Ded.",     "Standard"),
            ],
        ],
    })

    # Wind & Hail Deductible Buyback — row 36
    wh_carrier = _cell(s, "A36")
    coverages.append({
        "title": "Wind & Hail Deductible Buyback",
        "carrier": wh_carrier if not _is_placeholder(wh_carrier) else "Not Included",
        "not_included": _is_placeholder(wh_carrier),
        "description": (
            "A buyback policy reduces the wind and hail deductible on the "
            "primary property policy."
        ),
        "panel_rows": [[
            ("Deductible",        str(_cell(s, "D36") or "N/A")),
            ("Max Amount Payable", str(_cell(s, "F36") or "N/A")),
        ]],
    })

    # General Liability — row 43
    gl_carrier = _cell(s, "A43")
    coverages.append({
        "title": "Commercial General Liability",
        "carrier": gl_carrier if not _is_placeholder(gl_carrier) else "Not Included",
        "not_included": _is_placeholder(gl_carrier),
        "description": (
            "Commercial General Liability (CGL) protects the association "
            "against third-party claims for bodily injury, property damage, "
            "and personal and advertising injury in common areas."
        ),
        "panel_rows": [
            [
                ("Per Occurrence",         _money_str(_cell(s, "C43"))),
                ("General Aggregate",      _money_str(_cell(s, "I43"))),
                ("Products – Comp/Ops",    _money_str(_cell(s, "H43"))),
                ("Personal & Adv. Injury", _money_str(_cell(s, "G43"))),
            ],
            [
                ("Damage to Rented Premises", _money_str(_cell(s, "D43"))),
                ("Medical Expense",        _money_str(_cell(s, "F43"))),
                ("", ""),
                ("", ""),
            ],
        ],
    })

    # Hired & Non-Owned Auto — row 50
    auto_carrier = _cell(s, "A50")
    coverages.append({
        "title": "Hired & Non-Owned Auto (HNOA)",
        "carrier": auto_carrier if not _is_placeholder(auto_carrier) else "Not Included",
        "not_included": _is_placeholder(auto_carrier),
        "description": (
            "Covers liability for accidents involving vehicles the association "
            "uses for work purposes but does not own — including rentals and "
            "personal vehicles driven by board members, vendors, or volunteers."
        ),
        "panel_rows": [[
            ("Auto Symbols",           str(_cell(s, "C50") or "08 & 09")),
            ("Combined Single Limit",  f"{_money_str(_cell(s, 'D50'))} per accident"),
            ("Hired Auto Phys. Damage", "Not Included"),
            ("", ""),
        ]],
    })

    # Crime / Fidelity — row 57
    crime_carrier_cell = _cell(s, "A57")
    crime_limit = _cell(s, "C57")
    crime_carrier, crime_not_included = _coverage_status(
        crime_carrier_cell, [crime_limit], default_carrier=primary_carrier,
    )
    coverages.append({
        "title": "Crime / Fidelity",
        "carrier": crime_carrier,
        "not_included": crime_not_included,
        "description": (
            "Safeguards association funds against theft, fraud, or "
            "embezzlement by employees, volunteers, or agents. We strongly "
            "recommend this coverage to protect the board's fiduciary duty."
        ),
        "panel_rows": [[
            ("Employee Dishonesty Limit", _money_str(crime_limit, "Not Included")),
            ("Deductible",                _money_str(_cell(s, "G57"), "Standard")),
        ]],
    })

    # D&O — row 64
    do_carrier_cell = _cell(s, "A64")
    do_each = _cell(s, "C64")
    do_agg  = _cell(s, "F64")
    do_carrier, do_not_included = _coverage_status(
        do_carrier_cell, [do_each, do_agg], default_carrier=primary_carrier,
    )
    coverages.append({
        "title": "Directors & Officers Liability (D&O)",
        "carrier": do_carrier,
        "not_included": do_not_included,
        "description": (
            "Protects board members and officers from personal financial "
            "loss arising from association decisions. Covers legal defense, "
            "settlements, and damages. Intentional or willful misconduct is "
            "not covered."
        ),
        "panel_rows": [[
            ("Each Claim",     _money_str(do_each)),
            ("Aggregate",      _money_str(do_agg)),
            ("Cyber Liability", str(_cell(s, "I64") or "Not Included") if not _is_placeholder(_cell(s, "I64")) else "Not Included"),
            ("Retention",      str(_cell(s, "J64") or "Standard") if not _is_placeholder(_cell(s, "J64")) else "Standard"),
        ]],
    })

    # Umbrella — row 72
    umb_carrier = _cell(s, "A72")
    coverages.append({
        "title": "Umbrella \u2013 Excess Liability",
        "carrier": umb_carrier if not _is_placeholder(umb_carrier) else "Not Included",
        "not_included": _is_placeholder(umb_carrier),
        "description": (
            "Adds an extra layer of liability over the underlying GL, HNOA, "
            "and D&O policies."
        ),
        "panel_rows": [[
            ("Each Occurrence", _money_str(_cell(s, "C72"))),
            ("Aggregate",       _money_str(_cell(s, "F72"))),
            ("Retention",       _money_str(_cell(s, "I72"))),
            ("", ""),
        ]],
    })

    # Volunteer Accident — row 79
    va_carrier_cell = _cell(s, "A79")
    va_volunteers = _cell(s, "D79")
    va_limits_cell = _cell(s, "E79")
    va_volunteers_str = f"Up to {int(va_volunteers)}" if isinstance(va_volunteers, (int, float)) else str(va_volunteers or "—")
    va_carrier, va_not_included = _coverage_status(
        va_carrier_cell, [va_volunteers, va_limits_cell], default_carrier=primary_carrier,
    )
    coverages.append({
        "title": "Volunteer Accident",
        "carrier": va_carrier,
        "not_included": va_not_included,
        "description": (
            "Limited no-fault coverage for volunteers or participants injured "
            "while performing association activities. Benefits are payable "
            "up to the maximum stated limits."
        ),
        "panel_rows": [[
            ("# of Volunteers",  va_volunteers_str),
            ("AD&D / Paralysis / Medical", str(va_limits_cell or "—")),
            ("Deductible",       _money_str(_cell(s, "J79"), "$0")),
            ("", ""),
        ]],
    })

    # Workers' Comp — row 86
    wc_carrier = _cell(s, "A86")
    coverages.append({
        "title": "Workers' Compensation",
        "carrier": wc_carrier if not _is_placeholder(wc_carrier) else "Not Applicable",
        "not_included": _is_placeholder(wc_carrier),
        "description": (
            "Workers' Compensation covers employees injured on the job. In "
            "Texas, WC is not available for HOA volunteers, so it is "
            "typically not included on association proposals."
        ),
        "panel_rows": [[
            ("E.L. Each Accident",      _money_str(_cell(s, "C86"), "N/A")),
            ("E.L. Disease – Each Emp.", _money_str(_cell(s, "E86"), "N/A")),
            ("E.L. Disease – Policy",   _money_str(_cell(s, "G86"), "N/A")),
            ("# of Employees",          str(_cell(s, "I86") or "0 – If Any")),
        ]],
    })

    return coverages


def _parse_sov(wb) -> dict:
    """Read the SOV tab.

    The SOV tab varies dramatically across accounts. For now we read fixed
    section blocks; if a future template introduces other categories, this
    function is the place to extend.

    The current Haven at Keith Harrow Excel only has totals on the Summary
    tab; the SOV tab itself contains a long itemized list. To stay tolerant,
    we fall back to pulling values from the Summary tab's Property panel
    (rows 25-26) when the SOV tab is sparse.
    """
    summary = wb["Summary"]

    bldg_value    = _money(_cell(summary, "C26")) or 0.0
    bpp_value     = _money(_cell(summary, "D26")) or 0.0
    outdoor_value = _money(_cell(summary, "E26")) or 0.0
    total         = _money(_cell(summary, "G26")) or (bldg_value + bpp_value + outdoor_value)

    # Outdoor property breakdown — try SOV tab; if not present, use a sensible
    # default split. For Haven, the breakdown was 10,400 + 20,800 + 41,600 + 41,600.
    # Allocations like that aren't reliably in a known cell across templates,
    # so we present the rolled-up Outdoor Property line.
    outdoor_items = []
    if outdoor_value > 0:
        outdoor_items.append({
            "name": "Outdoor Property (consolidated)",
            "units": "—",
            "area": "—",
            "value": outdoor_value,
        })

    sections = [
        {
            "name": "Buildings",
            "items": [
                {"name": "Buildings — Non-Residential", "units": "—", "area": "—", "value": 0},
                {"name": "Buildings — Residential",     "units": "—", "area": "—", "value": bldg_value},
            ],
            "subtotal_label": "Total Buildings",
            "subtotal": bldg_value,
        },
        {
            "name": "Outdoor Property",
            "items": outdoor_items or [
                {"name": "Outdoor Property", "units": "—", "area": "—", "value": 0}
            ],
            "subtotal_label": "Total Outdoor Property",
            "subtotal": outdoor_value,
        },
        {
            "name": "Business Personal Property",
            "items": [
                {"name": "Business Personal Property", "units": "—", "area": "—", "value": bpp_value},
            ],
            "subtotal_label": "Total BPP",
            "subtotal": bpp_value,
        },
    ]

    return {"sections": sections, "total": total}


def _parse_authorization(wb) -> dict:
    auth = wb["Authorization"]

    # Premium-by-policy-type rows. Most cells in the current template are 0;
    # we only show non-zero or carrier-bearing lines so the page stays clean.
    raw_lines = [
        ("Commercial Package",          "G20"),
        ("Commercial Property",         "G22"),
        ("W/H Deductible Buyback",      "G24"),
        ("General Liability",           "G26"),
        ("Hired/Non-Owned Auto",        "G28"),
        ("Commercial Crime",            "G30"),
        ("Commercial D&O",          "G32"),
        ("Umbrella \u2013 Excess Liability", "G34"),
        ("Volunteer Accident",          "G36"),
        ("Workers' Compensation",       "G38"),
        ("Agency Fee",                  "G40"),
    ]
    policy_lines = []
    for label, coord in raw_lines:
        v = _money(_cell(auth, coord))
        if v is None or v == 0:
            continue
        policy_lines.append({"label": label, "proposed": _money_str(v)})

    total = _money(_cell(auth, "G42")) or 0.0

    pay_in_full = _money(_cell(auth, "G45")) or total
    down_payment = _money(_cell(auth, "F46"))
    install_amt  = _money(_cell(auth, "I46"))

    out = {
        "policy_lines": policy_lines,
        "total": total,
        "total_str": _money_str(total),
        "pay_in_full_str": _money_str(pay_in_full),
    }
    if down_payment is not None and install_amt is not None and down_payment > 0:
        out["down_payment_str"] = _money_str(down_payment)
        out["installment_amount_str"] = _money_str(install_amt)
        out["installments_count"] = 10
    return out


# -----------------------------------------------------------------------------
# Top-level
# -----------------------------------------------------------------------------
def parse_excel(file_obj_or_path) -> dict:
    """Parse a Blue Lime proposal Excel file.

    Args:
        file_obj_or_path: A file path, BytesIO, or any file-like object.

    Returns:
        A `ProposalData` dict ready for `proposal_generator.build_proposal()`.
    """
    if isinstance(file_obj_or_path, (bytes, bytearray)):
        file_obj_or_path = BytesIO(file_obj_or_path)

    wb = load_workbook(file_obj_or_path, data_only=True)

    return {
        "client":           _parse_client(wb),
        "account_manager":  _parse_account_manager(wb),
        "premium":          _parse_premium_comparison(wb),
        "coverages":        _parse_coverages(wb),
        "sov":              _parse_sov(wb),
        "authorization":    _parse_authorization(wb),
    }
